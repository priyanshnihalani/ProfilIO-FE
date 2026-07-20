from pathlib import Path

from openpyxl import load_workbook


FILES = [
    r"C:\Users\priya\Downloads\10th IB gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\4th IB southirving gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\6th IB west plano gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\8th IB gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\frisco 5th gen entries - matched highlighted.xlsx",
]


for file_name in FILES:
    path = Path(file_name)
    wb = load_workbook(path, data_only=False)
    ws = None
    for candidate in wb.worksheets:
        candidate_headers = [
            candidate.cell(1, col).value for col in range(1, candidate.max_column + 1)
        ]
        if "Matched Full Record(s)" in candidate_headers or "MatchFullRecord" in candidate_headers:
            ws = candidate
            break
    if ws is None:
        ws = wb.worksheets[0]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    print(f"FILE: {path}")
    print(f"SHEET: {ws.title} rows={ws.max_row} cols={ws.max_column}")
    print("HEADERS:", headers)
    match_col = 0
    for header in ("Matched Full Record(s)", "MatchFullRecord"):
        if header in headers:
            match_col = headers.index(header) + 1
            break
    print(f"MATCHCOL: {match_col}")
    if match_col:
        shown = 0
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, match_col).value
            if value not in (None, ""):
                print(f"ROW {row}: {repr(str(value)[:700])}")
                shown += 1
                if shown >= 3:
                    break
    print()
