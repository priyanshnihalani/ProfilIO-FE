from pathlib import Path

from openpyxl import load_workbook


OUTPUT_DIR = Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\matched-records-split")
FILES = sorted(OUTPUT_DIR.glob("*.xlsx"))
EXPECTED_HEADERS = [
    "Match 1 Store",
    "Match 1 Account No.",
    "Match 1 Bank Name",
    "Match 1 Post Date",
    "Match 1 Check No.",
    "Match 1 Description",
    "Match 1 Credit",
    "Match 1 Debit",
    "Match 1 Balance",
    "Match 1 Settle",
]


for path in FILES:
    wb = load_workbook(path, read_only=True, data_only=False)
    target = None
    for ws in wb.worksheets:
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        if EXPECTED_HEADERS[0] in headers:
            target = ws
            break
    if target is None:
        print(f"FAIL: {path.name} has no split headers")
        continue
    headers = [target.cell(1, col).value for col in range(1, target.max_column + 1)]
    start = headers.index(EXPECTED_HEADERS[0]) + 1
    found = headers[start - 1 : start - 1 + len(EXPECTED_HEADERS)]
    sample = None
    for row in range(2, target.max_row + 1):
        values = [target.cell(row, start + offset).value for offset in range(len(EXPECTED_HEADERS))]
        if any(value not in (None, "") for value in values):
            sample = (row, values)
            break
    match_groups = sum(1 for header in headers if isinstance(header, str) and header.endswith(" Store") and header.startswith("Match "))
    print(f"OK: {path.name} sheet={target.title} rows={target.max_row} match_groups={match_groups} split_headers={found == EXPECTED_HEADERS}")
    if sample:
        print(f"  SAMPLE R{sample[0]}: {sample[1]}")
