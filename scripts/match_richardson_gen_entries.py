from __future__ import annotations

from copy import copy
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT = Path(r"C:\Users\priya\OneDrive\Desktop\6th IB west plano gen entries.xlsx")
OUTPUT_DIR = Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\6th-ib-west-plano-gen-entries")
OUTPUT = OUTPUT_DIR / "6th IB west plano gen entries - matched highlighted.xlsx"

MATCH_FILL = PatternFill("solid", fgColor="C6EFCE")
MATCH_HEADER_FILL = PatternFill("solid", fgColor="70AD47")
DETAIL_HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")


def normalize_header(value: object) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def amount_key(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        text = str(value).replace("$", "").replace(",", "").strip()
        if not text:
            return None
        return Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def month_key(value: object) -> tuple[int, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if value is None or value == "":
        return None

    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.year, dt.month
        except ValueError:
            pass
    return None


def find_column(headers: list[object], candidates: set[str]) -> int:
    normalized = [normalize_header(h) for h in headers]
    for idx, header in enumerate(normalized, start=1):
        if header in candidates:
            return idx
    raise ValueError(f"Could not find any of {sorted(candidates)} in headers: {headers}")


def row_values(ws, row_idx: int, max_col: int) -> list[object]:
    return [ws.cell(row_idx, col_idx).value for col_idx in range(1, max_col + 1)]


def format_record(headers: list[object], values: list[object]) -> str:
    parts = []
    for header, value in zip(headers, values):
        if value is None or value == "":
            continue
        if isinstance(value, datetime):
            value = value.strftime("%Y-%m-%d")
        parts.append(f"{header}: {value}")
    return "; ".join(parts)


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def style_header(cell, fill) -> None:
    cell.fill = fill
    cell.font = Font(bold=True, color="FFFFFF")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(INPUT)

    sheet1 = wb["sheet1"]
    source_sheets = [wb[name] for name in ("Sheet2", "Sheet3", "Sheet4")]

    sheet1_max_col = sheet1.max_column
    sheet1_headers = row_values(sheet1, 1, sheet1_max_col)
    txn_date_col = find_column(sheet1_headers, {"txndate", "transactiondate", "date"})
    amount_col = find_column(sheet1_headers, {"amount", "txnamount", "transactionamount"})

    matches_by_key: dict[tuple[tuple[int, int], Decimal], list[dict[str, object]]] = {}
    source_matched_rows: dict[str, set[int]] = {ws.title: set() for ws in source_sheets}

    for ws in source_sheets:
        headers = row_values(ws, 1, ws.max_column)
        post_date_col = find_column(headers, {"postdate", "date"})
        debit_col = find_column(headers, {"debit", "debitamount"})

        for row_idx in range(2, ws.max_row + 1):
            post_key = month_key(ws.cell(row_idx, post_date_col).value)
            debit_key = amount_key(ws.cell(row_idx, debit_col).value)
            if post_key is None or debit_key is None or debit_key == 0:
                continue

            values = row_values(ws, row_idx, ws.max_column)
            entry = {
                "sheet": ws.title,
                "row": row_idx,
                "post_date": ws.cell(row_idx, post_date_col).value,
                "debit": ws.cell(row_idx, debit_col).value,
                "headers": headers,
                "values": values,
                "record": format_record(headers, values),
            }
            matches_by_key.setdefault((post_key, debit_key), []).append(entry)

    appended_headers = [
        "Match Status",
        "Match Count",
        "Matched Sheet(s)",
        "Matched Row(s)",
        "Matched Post Date(s)",
        "Matched Debit Amount(s)",
        "Matched Full Record(s)",
    ]
    append_start_col = sheet1_max_col + 1
    for offset, header in enumerate(appended_headers):
        cell = sheet1.cell(1, append_start_col + offset, header)
        style_header(cell, MATCH_HEADER_FILL)

    detail_name = "Matched Records Detail"
    if detail_name in wb.sheetnames:
        del wb[detail_name]
    detail = wb.create_sheet(detail_name)
    detail_headers = (
        ["Sheet1 Row"]
        + [f"Sheet1 {header}" for header in sheet1_headers]
        + ["Matched Source Sheet", "Matched Source Row", "Matched Full Source Record"]
    )
    max_source_cols = max(ws.max_column for ws in source_sheets)
    source_headers = []
    for idx in range(1, max_source_cols + 1):
        source_headers.append(f"Source Col {idx}")
    detail_headers += source_headers

    for col_idx, header in enumerate(detail_headers, start=1):
        cell = detail.cell(1, col_idx, header)
        style_header(cell, DETAIL_HEADER_FILL)

    detail_row = 2
    matched_sheet1_rows = 0

    for row_idx in range(2, sheet1.max_row + 1):
        key_month = month_key(sheet1.cell(row_idx, txn_date_col).value)
        key_amount = amount_key(sheet1.cell(row_idx, amount_col).value)
        row_matches = matches_by_key.get((key_month, key_amount), []) if key_month and key_amount is not None else []

        if not row_matches:
            sheet1.cell(row_idx, append_start_col, "No Match")
            sheet1.cell(row_idx, append_start_col + 1, 0)
            continue

        matched_sheet1_rows += 1
        for col_idx in range(1, append_start_col + len(appended_headers)):
            cell = sheet1.cell(row_idx, col_idx)
            if col_idx <= sheet1_max_col:
                copy_cell_style(sheet1.cell(row_idx, col_idx), cell)
            cell.fill = MATCH_FILL

        for match in row_matches:
            source_matched_rows[str(match["sheet"])].add(int(match["row"]))

        sheet_names = [str(m["sheet"]) for m in row_matches]
        row_numbers = [str(m["row"]) for m in row_matches]
        post_dates = [
            m["post_date"].strftime("%Y-%m-%d") if isinstance(m["post_date"], datetime) else str(m["post_date"])
            for m in row_matches
        ]
        debit_amounts = [str(m["debit"]) for m in row_matches]
        records = [str(m["record"]) for m in row_matches]
        summary_values = [
            "Matched",
            len(row_matches),
            ", ".join(sheet_names),
            ", ".join(row_numbers),
            ", ".join(post_dates),
            ", ".join(debit_amounts),
            " || ".join(records),
        ]
        for offset, value in enumerate(summary_values):
            sheet1.cell(row_idx, append_start_col + offset, value)

        sheet1_values = row_values(sheet1, row_idx, sheet1_max_col)
        for match in row_matches:
            values = (
                [row_idx]
                + sheet1_values
                + [match["sheet"], match["row"], match["record"]]
                + list(match["values"])
            )
            for col_idx, value in enumerate(values, start=1):
                detail.cell(detail_row, col_idx, value)
            detail_row += 1

    for ws in source_sheets:
        for row_idx in source_matched_rows[ws.title]:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row_idx, col_idx).fill = MATCH_FILL

    for ws in [sheet1, detail, *source_sheets]:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    width_overrides = {
        sheet1.title: {append_start_col + 6: 90},
        detail.title: {len(sheet1_headers) + 4: 90},
    }
    for ws in [sheet1, detail]:
        for col_idx in range(1, ws.max_column + 1):
            width = 14
            if ws.title in width_overrides and col_idx in width_overrides[ws.title]:
                width = width_overrides[ws.title][col_idx]
            elif col_idx <= ws.max_column:
                sample = [ws.cell(row, col_idx).value for row in range(1, min(ws.max_row, 50) + 1)]
                width = min(max(max((len(str(v)) for v in sample if v is not None), default=8) + 2, 10), 32)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(OUTPUT)
    print(f"Output: {OUTPUT}")
    print(f"Sheet1 matched rows: {matched_sheet1_rows}")
    print(f"Detail match rows: {detail_row - 2}")
    for name, rows in source_matched_rows.items():
        print(f"{name} matched rows: {len(rows)}")


if __name__ == "__main__":
    main()
