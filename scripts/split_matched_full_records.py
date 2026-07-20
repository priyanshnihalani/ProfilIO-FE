from copy import copy
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


FILES = [
    r"C:\Users\priya\Downloads\10th IB gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\4th IB southirving gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\6th IB west plano gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\8th IB gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\frisco 5th gen entries - matched highlighted.xlsx",
]

OUTPUT_DIR = Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\matched-records-split")
FULL_RECORD_HEADERS = ("Matched Full Record(s)", "Matched Bank Full Record(s)", "MatchFullRecord")
FIELD_HEADERS = [
    "Store",
    "Account No.",
    "Bank Name",
    "Post Date",
    "Check No.",
    "Description",
    "Credit",
    "Debit",
    "Balance",
    "Settle",
]
FIELD_KEYS = [
    "Store",
    "Account No.",
    "Bank Name",
    "Post Date",
    "Check No.",
    "Description",
    "Credit",
    "Debit",
    "Balance",
    "Settle",
]


def find_sheet_and_column(workbook):
    for worksheet in workbook.worksheets:
        headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
        for header in FULL_RECORD_HEADERS:
            if header in headers:
                return worksheet, headers.index(header) + 1
    return None, None


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.border:
        target.border = copy(source.border)
    if source.fill:
        target.fill = copy(source.fill)
    if source.protection:
        target.protection = copy(source.protection)
    if source.font:
        target.font = copy(source.font)


def parse_record(value):
    if value in (None, ""):
        return {key: None for key in FIELD_KEYS}

    text = str(value).replace("\r\n", "\n").strip()
    result = {}
    for index, key in enumerate(FIELD_KEYS):
        next_keys = FIELD_KEYS[index + 1 :]
        if next_keys:
            lookahead = "|".join(re.escape(f"; {next_key}:") for next_key in next_keys)
            pattern = rf"{re.escape(key)}:\s*(.*?)(?={lookahead}|$)"
        else:
            pattern = rf"{re.escape(key)}:\s*(.*)$"
        match = re.search(pattern, text, flags=re.DOTALL)
        result[key] = match.group(1).strip() if match else None

    return result


def split_records(value):
    if value in (None, ""):
        return []
    return [part.strip() for part in str(value).split(" || ") if part.strip()]


def split_workbook(path):
    workbook = load_workbook(path)
    worksheet, full_record_col = find_sheet_and_column(workbook)
    if worksheet is None:
        return None

    max_records = 1
    for row in range(2, worksheet.max_row + 1):
        max_records = max(max_records, len(split_records(worksheet.cell(row, full_record_col).value)))

    insert_at = full_record_col + 1
    total_insert_cols = len(FIELD_HEADERS) * max_records
    worksheet.insert_cols(insert_at, total_insert_cols)

    header_source = worksheet.cell(1, full_record_col)
    data_source = worksheet.cell(2, full_record_col)
    output_headers = []
    for match_index in range(1, max_records + 1):
        output_headers.extend([f"Match {match_index} {header}" for header in FIELD_HEADERS])

    for offset, header in enumerate(output_headers):
        cell = worksheet.cell(1, insert_at + offset)
        cell.value = header
        copy_cell_style(header_source, cell)
        worksheet.column_dimensions[get_column_letter(insert_at + offset)].width = max(14, min(28, len(header) + 3))

    for row in range(2, worksheet.max_row + 1):
        records = split_records(worksheet.cell(row, full_record_col).value)
        for match_offset in range(max_records):
            parsed = parse_record(records[match_offset]) if match_offset < len(records) else {}
            for field_offset, key in enumerate(FIELD_KEYS):
                output_offset = match_offset * len(FIELD_KEYS) + field_offset
                cell = worksheet.cell(row, insert_at + output_offset)
                cell.value = parsed.get(key)
                copy_cell_style(data_source, cell)

    for match_offset in range(max_records):
        group_start = insert_at + match_offset * len(FIELD_HEADERS)
        worksheet.column_dimensions[get_column_letter(group_start + 5)].width = 48
        for col in (group_start + 3, group_start + 4):
            worksheet.column_dimensions[get_column_letter(col)].width = 16

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / path.name.replace(" - matched highlighted.xlsx", " - matched records split.xlsx")
    workbook.save(output_path)
    return output_path, worksheet.title, full_record_col, insert_at, max_records


if __name__ == "__main__":
    for file_name in FILES:
        input_path = Path(file_name)
        result = split_workbook(input_path)
        if result is None:
            print(f"SKIPPED: {input_path.name} - no matched full record column found")
            continue
        output_path, sheet_name, full_record_col, insert_at, max_records = result
        print(
            f"WROTE: {output_path} | sheet={sheet_name} | full_record_col={full_record_col} | split_start_col={insert_at} | max_matches={max_records}"
        )
