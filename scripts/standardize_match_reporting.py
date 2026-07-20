from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


FILES = [
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\richardson-gen-entries\richardson gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\frisco-5th-gen-entries\frisco 5th gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\ots400-gen-entries\OTS400 gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\4th-ib-southirving-gen-entries\4th IB southirving gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\ots200-gen-entries\OTS200 gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\ots100-gen-entries\OTS100 gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\plano-gen-entries\plano gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\10th-ib-gen-entries\10th IB gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\lewisville-gen-entries\lewisville gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\8th-ib-gen-entries\8th IB gen entries - matched highlighted.xlsx"),
    Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\6th-ib-west-plano-gen-entries\6th IB west plano gen entries - matched highlighted.xlsx"),
]

STANDARDIZED_DIR = Path(r"D:\AI_Integration\AI_Resume_Generator\resume_generator\outputs\standardized-match-reporting")


HEADER_RENAMES = {
    "Match Count": "Possible Bank Match Count",
    "Matched Sheet(s)": "Matched Bank Sheet(s)",
    "Matched Row(s)": "Matched Bank Row(s)",
    "Matched Post Date(s)": "Matched Bank Post Date(s)",
    "Matched Debit Amount(s)": "Matched Bank Debit Amount(s)",
    "Matched Full Record(s)": "Matched Bank Full Record(s)",
    "Matched Source Sheet": "Matched Bank Sheet",
    "Matched Source Row": "Matched Bank Row",
    "Matched Full Source Record": "Matched Bank Full Record",
}

REPORT_SHEET = "How to Read Matches"
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")


def normalize_header(value: object) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def find_journal_sheet(wb):
    for ws in wb.worksheets:
        headers = {normalize_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)}
        if {"txndate", "amount"}.issubset(headers):
            return ws
    raise ValueError("Journal sheet with txnDate and amount was not found")


def find_bank_sheets(wb):
    bank_sheets = []
    for ws in wb.worksheets:
        headers = {normalize_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)}
        if {"postdate", "debit"}.issubset(headers):
            bank_sheets.append(ws)
    return bank_sheets


def rename_headers(ws) -> None:
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value in HEADER_RENAMES:
            ws.cell(1, col).value = HEADER_RENAMES[value]


def put(ws, row: int, col: int, value: str, bold: bool = False) -> None:
    cell = ws.cell(row, col, value)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if bold:
        cell.font = Font(bold=True)


def create_report_sheet(wb, journal_sheet_name: str, bank_sheet_names: list[str]) -> None:
    if REPORT_SHEET in wb.sheetnames:
        del wb[REPORT_SHEET]

    ws = wb.create_sheet(REPORT_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "Match Reporting Guide"
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = TITLE_FILL
    title.alignment = Alignment(horizontal="center")

    for cell in ws[3]:
        cell.fill = SUBTITLE_FILL
        cell.font = Font(bold=True)

    rows = [
        ("Workbook status", "Completed and ready for review."),
        ("Matching rule", "Journal txnDate month/year + amount was compared with bank Post Date month/year + Debit."),
        ("Where to review journal matches", f"Open {journal_sheet_name}. Green rows are journal rows that found at least one possible bank debit match."),
        ("Where to review bank matches", f"Open {', '.join(bank_sheet_names)}. Green rows are bank debit rows that were used as possible matches."),
        ("Where to review full details", "Open Matched Records Detail. It shows the full journal row beside the full bank row for each possible match."),
        ("Important note", "Do not compare the row totals across sheets. Sheet rows and detail rows measure different things when the same amount repeats in the same month."),
        ("No-match rows", "Rows marked No Match did not find a bank Debit with the same month/year and amount."),
    ]

    put(ws, 3, 1, "Section", True)
    put(ws, 3, 2, "Explanation", True)

    for idx, (section, explanation) in enumerate(rows, start=4):
        put(ws, idx, 1, section, True)
        put(ws, idx, 2, explanation)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 4

    for row in range(1, len(rows) + 4):
        ws.row_dimensions[row].height = 32 if row >= 4 else 24


def verify_workbook(wb) -> tuple[str, list[str]]:
    journal = find_journal_sheet(wb)
    bank_sheets = find_bank_sheets(wb)
    if not bank_sheets:
        raise ValueError("No bank sheets with Post Date and Debit were found")
    if "Matched Records Detail" not in wb.sheetnames:
        raise ValueError("Matched Records Detail sheet was not found")

    headers = [journal.cell(1, col).value for col in range(1, journal.max_column + 1)]
    required = {
        "Match Status",
        "Possible Bank Match Count",
        "Matched Bank Sheet(s)",
        "Matched Bank Row(s)",
        "Matched Bank Post Date(s)",
        "Matched Bank Debit Amount(s)",
        "Matched Bank Full Record(s)",
    }
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"Missing journal report headers: {sorted(missing)}")

    return journal.title, [ws.title for ws in bank_sheets]


def main() -> None:
    STANDARDIZED_DIR.mkdir(parents=True, exist_ok=True)

    for path in FILES:
        if not path.exists():
            raise FileNotFoundError(path)

        wb = load_workbook(path)
        for ws in wb.worksheets:
            rename_headers(ws)

        journal_name, bank_names = verify_workbook(wb)
        create_report_sheet(wb, journal_name, bank_names)

        for ws in wb.worksheets:
            if ws.max_column:
                ws.auto_filter.ref = ws.dimensions

        output_path = STANDARDIZED_DIR / path.name
        wb.save(output_path)
        print(f"Saved standardized copy: {output_path.name}")

    print(f"Done: {len(FILES)} standardized workbooks saved")


if __name__ == "__main__":
    main()
