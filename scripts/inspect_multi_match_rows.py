from pathlib import Path

from openpyxl import load_workbook


FILES = [
    r"C:\Users\priya\Downloads\10th IB gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\4th IB southirving gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\6th IB west plano gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\8th IB gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\frisco 5th gen entries - matched highlighted.xlsx",
]


for file_name in FILES:
    path = Path(file_name)
    wb = load_workbook(path, read_only=True, data_only=False)
    print(f"FILE: {path.name}")
    for ws in wb.worksheets:
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        record_names = ("Matched Full Record(s)", "Matched Bank Full Record(s)", "MatchFullRecord")
        count_names = ("Match Count", "Possible Bank Match Count")
        record_col = next((headers.index(name) + 1 for name in record_names if name in headers), 0)
        count_col = next((headers.index(name) + 1 for name in count_names if name in headers), 0)
        if not record_col or not count_col:
            continue
        shown = 0
        for row in range(2, ws.max_row + 1):
            count = ws.cell(row, count_col).value
            try:
                count_number = int(count)
            except (TypeError, ValueError):
                count_number = 0
            if count_number > 1:
                print(f"  {ws.title} R{row} count={count_number}: {repr(str(ws.cell(row, record_col).value)[:1000])}")
                shown += 1
                if shown >= 3:
                    break
        if shown == 0:
            print(f"  {ws.title}: no rows with match count > 1")
