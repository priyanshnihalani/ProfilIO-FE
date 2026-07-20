from openpyxl import load_workbook


FILES = [
    r"C:\Users\priya\Downloads\4th IB southirving gen entries - matched highlighted.xlsx",
    r"C:\Users\priya\Downloads\frisco 5th gen entries - matched highlighted.xlsx",
]


for file_name in FILES:
    print(f"FILE: {file_name}")
    wb = load_workbook(file_name, read_only=True, data_only=False)
    for ws in wb.worksheets:
        print(f"SHEET: {ws.title} rows={ws.max_row} cols={ws.max_column}")
        for row in range(1, min(ws.max_row, 5) + 1):
            values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 25) + 1)]
            print(f"  R{row}: {values}")
    print()
